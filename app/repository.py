"""المستودع المعرفي — تحليل ملفات العروض القديمة (لعزوم أو لشركات أخرى)
واستخراج بنودها المسعّرة كأسعار سوق استرشادية، وتخزين نصوصها لتغذية التوليد.
"""
import re
from io import BytesIO
from pathlib import Path

from .file_extract import OCR_MARKER, extract_text

# كلمات رؤوس الأعمدة الشائعة في جداول الكميات — مرتبة بالأولوية (الوصف قبل رقم البند)
_HDR_NAME = ("وصف", "بيان", "description", "الاعمال", "الأعمال", "البند", "item")
_HDR_UNIT = ("وحدة", "الوحدة", "unit")
_HDR_QTY = ("كمية", "الكمية", "qty", "quantity")
_HDR_RATE = ("سعر", "السعر", "افرادى", "إفرادى", "rate", "price", "فئة")


def _num(v):
    try:
        x = float(str(v).replace(",", "").strip())
        return x if x == x else None  # استبعاد NaN
    except (ValueError, TypeError):
        return None


def _clean(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _hdr_norm(s) -> str:
    """تطبيع نص الرأس: إزالة المدّات (ـ) والمسافات لمطابقة «الـــوصــــف» مع «وصف»."""
    return re.sub(r"[ـ\s]+", "", str(s or "")).lower()


def _match_col(header_cells: list, keywords: tuple) -> list[int]:
    """الأعمدة المطابقة مرتبة بأولوية الكلمات المفتاحية ثم بموقع العمود."""
    out = []
    for prio, k in enumerate(keywords):
        for i, c in enumerate(header_cells):
            text = _hdr_norm(c)
            if text and k in text and i not in out:
                out.append(i)
    return out


# تحويل الأرقام العربية والفواصل العربية (٫ عشرية، ٬ فاصل آلاف)
_AR_TRANS = str.maketrans("٠١٢٣٤٥٦٧٨٩٫٬", "0123456789.,")
_NUM_RE = re.compile(r"\d+(?:[,،]\d{3})*(?:\.\d+)?")
_UNITS = ("م2", "م3", "م²", "م³", "م.ط", "مط", "متر مربع", "متر طولي", "متر مكعب",
          "عدد", "بالعدد", "مقطوعية", "لس", "شهر", "سنة", "يوم", "نقطة", "طن",
          "m2", "m3", "m²", "no", "nos", "ls", "item", "mrun", "lm", "kg")
_TOTAL_WORDS = ("إجمالي", "الاجمالي", "الإجمالي", "مجموع", "total", "subtotal", "الاجمالى")


def parse_text_boq(text: str) -> list[dict]:
    """استخراج البنود المسعّرة من نص مستخرج (PDF/Word/نص) — سطر يحمل وصفاً
    وثلاثة أرقام تحقق الكمية × سعر الوحدة = الإجمالي (بأي اتجاه قراءة).

    عند استخراج PDF كثيراً ما يلتفّ وصف البند على سطر مستقل عن سطر الأرقام،
    لذا يُحتفظ بآخر سطر نصي كوصف احتياطي لسطر الأرقام التالي."""
    items, seen = [], set()
    pending_desc = ""
    for raw_line in text.splitlines():
        line = raw_line.translate(_AR_TRANS)
        matches = _NUM_RE.findall(line)
        if len(matches) < 3:
            # سطر نصي — قد يكون وصف بند التفّ قبل سطر أرقامه
            stripped = _clean(_NUM_RE.sub(" ", line).replace("|", " "))
            stripped = re.sub(r"^[\W_]+|[\W_]+$", "", stripped)
            if len(stripped) >= 8 and not any(w in stripped for w in _TOTAL_WORDS):
                pending_desc = stripped
            continue
        nums = []
        for m in matches:
            try:
                nums.append(float(m.replace(",", "").replace("،", "")))
            except ValueError:
                pass
        desc = _clean(_NUM_RE.sub(" ", line).replace("|", " ").replace("/", " "))
        desc = re.sub(r"^[\W_]+|[\W_]+$", "", desc)
        if len(desc) < 8:
            desc = pending_desc  # إنقاذ الوصف من السطر السابق
        if len(desc) < 8 or any(w in desc for w in _TOTAL_WORDS):
            continue
        qty = rate = None
        for i in range(len(nums) - 2):
            a, b, c = nums[i], nums[i + 1], nums[i + 2]
            # الترتيب الطبيعي: كمية × سعر = إجمالي
            if a > 0 and b > 0 and c > 0 and abs(a * b - c) <= max(1.0, 0.02 * c):
                qty, rate = a, b
                break
            # الترتيب المعكوس (استخراج RTL): إجمالي، سعر، كمية
            if abs(c * b - a) <= max(1.0, 0.02 * a) and c > 0 and b > 0 and a > 0:
                qty, rate = c, b
                break
        if rate is None or rate <= 0 or rate > 50_000_000:
            continue
        unit = next((u for u in _UNITS if u in raw_line or u in line.lower()), "")
        key = (desc[:60], rate)
        if key in seen:
            continue
        seen.add(key)
        items.append({"name": desc[:200], "unit": unit, "qty": qty, "unit_price": rate})
        pending_desc = ""
    return items


def parse_boq_items(content: bytes, filename: str) -> list[dict]:
    """استخراج البنود المسعّرة: Excel باكتشاف الأعمدة، وPDF/Word/نص بمحلل الأسطر."""
    if Path(filename).suffix.lower() not in (".xlsx", ".xlsm"):
        return parse_text_boq(extract_text(filename, content))
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return []

    items = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        cols = _detect_columns(rows)
        if not cols:
            continue
        name_c, unit_c, qty_c, rate_c, start = cols
        for row in rows[start:]:
            if len(row) <= max(name_c, rate_c):
                continue
            name = _clean(row[name_c])
            rate = _num(row[rate_c])
            if not name or len(name) < 6 or not rate or rate <= 0:
                continue
            # استبعاد صفوف الإجماليات
            if any(k in name for k in ("إجمالي", "الاجمالي", "total", "Total", "TOTAL", "مجموع")):
                continue
            items.append({
                "name": name[:200],
                "unit": _clean(row[unit_c])[:20] if unit_c is not None and len(row) > unit_c else "",
                "qty": _num(row[qty_c]) if qty_c is not None and len(row) > qty_c else None,
                "unit_price": rate,
            })
    # لم تُكتشف رؤوس أعمدة (تنسيق غير معتاد) — جرّب محلل الأسطر على النص المستخرج
    if not items:
        items = parse_text_boq(extract_text(filename, content))
    return items


def _detect_columns(rows) -> tuple | None:
    """البحث عن صف الرؤوس في أول 15 صفاً وتحديد مواقع الأعمدة."""
    for idx, row in enumerate(rows[:15]):
        cells = ["" if c is None else str(c) for c in row]
        names = _match_col(cells, _HDR_NAME)
        rates = _match_col(cells, _HDR_RATE)
        if not names or not rates:
            continue
        units = _match_col(cells, _HDR_UNIT)
        qtys = _match_col(cells, _HDR_QTY)
        # استبعاد أعمدة "الإجمالي" من مرشحي السعر (سعر الوحدة لا الإجمالي)
        rate_c = None
        for r in rates:
            h = _hdr_norm(cells[r])
            if "جمال" not in h and "amount" not in h and "total" not in h:
                rate_c = r
                break
        if rate_c is None:
            continue
        # أول عمود وصف بالأولوية، مع تفضيل العربي بين أعمدة الوصف فقط (لا أعمدة رقم البند)
        name_c = names[0]
        desc_cols = [n for n in names if any(k in _hdr_norm(cells[n]) for k in ("وصف", "بيان", "description"))]
        for n in desc_cols:
            if any("؀" <= ch <= "ۿ" for ch in cells[n]):
                name_c = n
                break
        else:
            if desc_cols:
                name_c = desc_cols[0]
        return (name_c, units[0] if units else None, qtys[0] if qtys else None, rate_c, idx + 1)
    return None


def find_relevant_repo_texts(query: str, top_n: int = 2) -> list[dict]:
    """أقرب ملفات المستودع لنص المشروع — تُضاف كسياق معرفي عند التوليد."""
    from .database import get_repo_texts
    from .similarity import _tokens
    q = _tokens(query)
    if not q:
        return []
    scored = []
    for f in get_repo_texts(limit_chars=200_000):
        overlap = len(q & _tokens(f["extracted_text"] + " " + f["filename"]))
        if overlap >= 3:
            # نمرر مقتطفاً فقط للسياق حتى لا نضخم طلب التوليد
            scored.append((overlap, {**f, "extracted_text": f["extracted_text"][:6000]}))
    scored.sort(key=lambda x: -x[0])
    return [f for _, f in scored[:top_n]]


_SCOPE_HINTS = ("أعمال", "اعمال", "توريد", "تركيب", "تنفيذ", "إنشاء", "انشاء",
                "صيانة", "تشغيل", "تجهيز", "تأهيل", "ترميم", "عزل", "تشطيب")


def build_reference_from_text(filename: str, company: str, text: str, items: list[dict]) -> dict | None:
    """تحويل ملف عرض فني/مالي محلل إلى عرض مرجعي في ذاكرة النظام —
    يستخرج العنوان والملخص والنطاق والكلمات المفتاحية وجدول الكميات،
    فيدخل العرض في محرك التشابه ويُبنى عليه في المشاريع المشابهة."""
    from collections import Counter
    from .database import create_proposal, list_proposals_full
    from .similarity import _tokens

    clean_text = text.replace(OCR_MARKER, "") if text else ""
    if len(re.sub(r"\s", "", clean_text)) < 150 and not items:
        return None  # لا محتوى يُبنى عليه

    title = re.sub(r"[_\-]+", " ", Path(filename).stem)
    title = _clean(title)[:120] or "عرض مرجعي"
    existing = {p["title"] for p in list_proposals_full() if p["data"].get("reference")}
    if title in existing:
        return {"duplicate": True, "title": title}

    lines = [l.strip() for l in clean_text.splitlines() if len(l.strip()) > 15]
    summary = _clean(" ".join(lines[:6]))[:800]
    scope = []
    for l in lines:
        if len(scope) >= 12:
            break
        looks_bullet = bool(re.match(r"^[-•*–]|^\d{1,2}[\).\-]", l))
        if (looks_bullet or any(k in l for k in _SCOPE_HINTS)) and len(l) < 200:
            item = _clean(re.sub(r"^[-•*–\d\).\s]+", "", l))[:160]
            if len(item) > 10 and item not in scope:
                scope.append(item)
    freq = Counter()
    for l in lines[:400]:
        freq.update(_tokens(l))
    keywords = " ".join(w for w, _ in freq.most_common(15))

    boq = []
    for it in items[:300]:
        qty = it.get("qty") or 1
        boq.append({
            "code": "", "name": it["name"], "unit": it.get("unit", ""),
            "qty": qty, "unit_price": it["unit_price"],
            "total": round(qty * it["unit_price"], 2),
        })
    direct = round(sum(l["total"] for l in boq), 2)
    data = {
        "reference": True,
        "summary": summary,
        "scope": scope,
        "keywords": keywords,
        "technical_sections": [],
        "boq": boq,
        "financial": {"direct_cost": direct} if direct else {},
        "plan": [],
        "engine": "المستودع المعرفي — ملف محلل",
        "source_file": filename,
    }
    p = create_proposal(title, company or "غير محدد", "government", data)
    return {"id": p["id"], "ref_no": p["ref_no"], "title": title, "boq_lines": len(boq)}


def ingest_file(filename: str, content: bytes, source_type: str, company: str,
                notes: str, as_reference: bool = False) -> dict:
    """تحليل ملف مرفوع وتخزينه في المستودع: نص كامل + بنود مسعّرة + تشخيص،
    مع خيار إضافته كعرض مرجعي في ذاكرة التشابه."""
    from .database import create_repo_file
    from .file_extract import ocr_available
    text = extract_text(filename, content)
    if Path(filename).suffix.lower() in (".xlsx", ".xlsm"):
        items = parse_boq_items(content, filename)
    else:
        items = parse_text_boq(text)
    meta = {"filename": filename, "source_type": source_type, "company": company, "notes": notes}
    record = create_repo_file(meta, text, items)

    if as_reference:
        ref = build_reference_from_text(filename, company, text, items)
        if ref is None:
            record["reference_note"] = "تعذر بناء عرض مرجعي — لا يوجد محتوى نصي كافٍ في الملف."
        elif ref.get("duplicate"):
            record["reference_note"] = f"يوجد عرض مرجعي بنفس العنوان مسبقاً: {ref['title']}"
        else:
            record["reference"] = ref

    # تشخيص واضح عندما لا تُستخرج بنود
    if not items:
        clean_len = len(re.sub(r"\s", "", text))
        if text.startswith("[تعذر") or clean_len < 150:
            if ocr_available():
                record["note"] = ("لم يُعثر على نص قابل للقراءة حتى بعد محاولة الاستخراج الضوئي — "
                                  "جودة المسح منخفضة. ارفع نسخة Excel أو PDF نصياً.")
            else:
                record["note"] = ("الملف مصور (سكان) وأدوات الاستخراج الضوئي OCR غير مثبتة على الخادم. "
                                  "ثبّتها بالأمر: apt install tesseract-ocr tesseract-ocr-ara poppler-utils "
                                  "ثم أعد رفع الملف — أو ارفع نسخة Excel/PDF نصية.")
        else:
            record["note"] = ("خُزّن النص للاستفادة المعرفية، لكن لم تُرصد أسطر مسعّرة "
                              "(كمية × سعر = إجمالي). إن كان الملف يحوي أسعاراً بصيغة مختلفة أرسله لنا لضبط المحلل.")
        if text.startswith(OCR_MARKER):
            record["note"] = "قُرئ الملف بالاستخراج الضوئي OCR. " + record.get("note", "")
    elif text.startswith(OCR_MARKER):
        record["note"] = f"قُرئ الملف المصور بالاستخراج الضوئي OCR واستُخرج {len(items)} بنداً."
    return record
